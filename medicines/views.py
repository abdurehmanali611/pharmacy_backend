from django.utils.dateparse import parse_date
from rest_framework import permissions, viewsets
from rest_framework.decorators import action
from rest_framework.exceptions import ValidationError
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.parsers import MultiPartParser
from rest_framework.response import Response
from rest_framework import status

from openpyxl import load_workbook
from suppliers.models import Supplier

from .models import Medicine
from .serializers import MedicineSerializer

class MedicineViewSet(viewsets.ModelViewSet):
    queryset = Medicine.objects.all()
    serializer_class = MedicineSerializer
    permission_classes = [permissions.IsAuthenticated]
    filter_backends = [SearchFilter, OrderingFilter]
    search_fields = ["name", "category", "supplier_name", "supplier_phone", "supplier_email", "description"]
    ordering_fields = ["created_at", "updated_at", "name", "category", "quantity", "price", "cost", "expiry_date"]
    ordering = ["-created_at"]

    def get_queryset(self):
        profile = getattr(self.request.user, "profile", None)
        pharmacy_tin = getattr(profile, "pharmacy_tin", "")
        if not pharmacy_tin:
            return Medicine.objects.none()
        qs = super().get_queryset().filter(pharmacy_tin=pharmacy_tin)

        # Optional filter by day the medicine was registered/created.
        created_date = self.request.query_params.get("created_date")
        if created_date:
            parsed = parse_date(created_date)
            if not parsed:
                raise ValidationError({"created_date": ["Invalid date format. Use YYYY-MM-DD."]})
            qs = qs.filter(created_at__date=parsed)

        return qs
    
    def perform_create(self, serializer):
        profile = getattr(self.request.user, "profile", None)
        pharmacy_tin = getattr(profile, "pharmacy_tin", "")
        if not pharmacy_tin:
            raise ValidationError({"detail": "Missing pharmacy TIN for the current user."})
        serializer.save(pharmacy_tin=pharmacy_tin)

    def perform_update(self, serializer):
        profile = getattr(self.request.user, "profile", None)
        pharmacy_tin = getattr(profile, "pharmacy_tin", "")
        if not pharmacy_tin:
            raise ValidationError({"detail": "Missing pharmacy TIN for the current user."})
        serializer.save(pharmacy_tin=pharmacy_tin)

    @action(detail=False, methods=["post"], url_path="import_excel", parser_classes=[MultiPartParser])
    def import_excel(self, request):
        """
        Accepts an .xlsx file and bulk-creates medicines for the current pharmacy.

        Expected columns (header row, case-insensitive):
        - name (required)
        - category (optional)
        - price (required)
        - cost (optional)
        - quantity (required)
        - batch_number (optional)
        - expiry_date (optional, YYYY-MM-DD or Excel date)
        - description (required)
        - supplier_name (required)
        - supplier_phone (required)
        - supplier_email (optional)
        """
        profile = getattr(request.user, "profile", None)
        pharmacy_tin = getattr(profile, "pharmacy_tin", "")
        if not pharmacy_tin:
            raise ValidationError({"detail": "Missing pharmacy TIN for the current user."})

        upload = request.FILES.get("file")
        if not upload:
            raise ValidationError({"file": ["Missing file."]})
        if not upload.name.lower().endswith(".xlsx"):
            raise ValidationError({"file": ["Only .xlsx files are supported."]})

        try:
            wb = load_workbook(filename=upload, data_only=True)
            ws = wb.active
        except Exception:
            raise ValidationError({"file": ["Invalid Excel file."]})

        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            raise ValidationError({"file": ["Excel file is empty."]})

        header = [str(c).strip().lower() if c is not None else "" for c in rows[0]]
        if not any(header):
            raise ValidationError({"file": ["Missing header row."]})

        def col(name: str):
            return header.index(name) if name in header else None

        required = ["name", "price", "quantity", "description", "supplier_name", "supplier_phone"]
        missing = [c for c in required if c not in header]
        if missing:
            raise ValidationError({"file": [f"Missing required columns: {', '.join(missing)}"]})

        idx = {
            "name": col("name"),
            "category": col("category"),
            "price": col("price"),
            "cost": col("cost"),
            "quantity": col("quantity"),
            "batch_number": col("batch_number"),
            "expiry_date": col("expiry_date"),
            "description": col("description"),
            "supplier_name": col("supplier_name"),
            "supplier_phone": col("supplier_phone"),
            "supplier_email": col("supplier_email"),
        }

        created = 0
        errors = []

        for i, row in enumerate(rows[1:], start=2):  # 1-based row index, header=1
            # skip empty rows
            if not row or all(cell is None or str(cell).strip() == "" for cell in row):
                continue

            def get(key: str):
                j = idx.get(key)
                return row[j] if j is not None and j < len(row) else None

            payload = {
                "name": (str(get("name")).strip() if get("name") is not None else ""),
                "category": (str(get("category")).strip() if get("category") is not None else ""),
                "price": get("price"),
                "cost": get("cost") if get("cost") is not None else 0,
                "quantity": get("quantity"),
                "batch_number": (str(get("batch_number")).strip() if get("batch_number") is not None else ""),
                "expiry_date": get("expiry_date"),
                "description": (str(get("description")).strip() if get("description") is not None else ""),
                "supplier_name": (str(get("supplier_name")).strip() if get("supplier_name") is not None else ""),
                "supplier_phone": (str(get("supplier_phone")).strip() if get("supplier_phone") is not None else ""),
                "supplier_email": (str(get("supplier_email")).strip() if get("supplier_email") is not None else ""),
            }

            # Normalize expiry_date
            expiry = payload.get("expiry_date")
            if isinstance(expiry, str):
                expiry = parse_date(expiry.strip())
            # If Excel date comes as datetime/date, keep only date.
            if hasattr(expiry, "date"):
                try:
                    expiry = expiry.date()
                except Exception:
                    pass
            payload["expiry_date"] = expiry if expiry else None

            supplier_name = payload["supplier_name"]
            supplier = None
            if supplier_name:
                supplier, _ = Supplier.objects.get_or_create(
                    pharmacy_tin=pharmacy_tin,
                    supplier_name=supplier_name,
                    defaults={
                        "supplier_phone": payload["supplier_phone"],
                        "supplier_email": payload["supplier_email"],
                    },
                )

                supplier_changed = False
                if payload["supplier_phone"] and supplier.supplier_phone != payload["supplier_phone"]:
                    supplier.supplier_phone = payload["supplier_phone"]
                    supplier_changed = True
                if payload["supplier_email"] != supplier.supplier_email:
                    supplier.supplier_email = payload["supplier_email"]
                    supplier_changed = True
                if supplier_changed:
                    supplier.save(update_fields=["supplier_phone", "supplier_email", "updated_at"])

            serializer_payload = {
                **payload,
                "supplier_id": supplier.id if supplier else None,
            }

            serializer = MedicineSerializer(data=serializer_payload, context={"request": request})
            if not serializer.is_valid():
                errors.append({"row": i, "errors": serializer.errors})
                continue

            serializer.save(pharmacy_tin=pharmacy_tin)
            created += 1

        return Response(
            {
                "detail": f"Imported {created} medicines.",
                "created": created,
                "failed": len(errors),
                "errors": errors[:50],
            },
            status=status.HTTP_200_OK,
        )
